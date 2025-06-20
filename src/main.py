from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
from flask_cors import CORS
import os
import sys
import atexit
import pandas as pd
import tempfile
from datetime import datetime
import logging
from src.book_sites.open_library import BookScraper
from src.book_sites.barnes_and_noble import BarnesNobleBookScraper
from src.book_sites.thriftbooks import ThriftBooksBookScraper

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Updated Flask configuration to match your file structure
app = Flask(__name__, static_folder='static', template_folder='template')
CORS(app)

# Initialize the book scraper
scraper = BookScraper()
barnes_noble_scraper = BarnesNobleBookScraper()
thrift_books_scraper = ThriftBooksBookScraper()

# Store search results temporarily (in production, use a proper database)
search_cache = {}

@app.route('/')
def index():
    """Serve the main HTML page"""
    return render_template('index.html')

@app.route('/style.css')
def styles():
    """Serve the CSS file"""
    return send_from_directory('src/static', 'style.css')

@app.route('/scripts.js')
def script():
    """Serve the JavaScript file"""
    return send_from_directory('src/static', 'scripts.js')

@app.route('/api/search', methods=['POST'])
def search_books():
    """Handle book search requests from the frontend"""
    try:
        data = request.get_json()
        
        # Extract search parameters
        book_name = data.get('bookName', '').strip()
        isbn = data.get('isbn', '').strip()
        author = data.get('author', '').strip()
        site = data.get('site', '')
        
        # Validate input
        if not any([book_name, isbn, author]):
            return jsonify({
                'success': False,
                'message': 'Please provide at least one search parameter (book name, ISBN, or author)'
            }), 400
        
        # Determine search query priority: ISBN > Book Name > Author
        search_query = isbn if isbn else (book_name if book_name else author)
        search_type = 'isbn' if isbn else ('title' if book_name else 'author')
        
        logger.info(f"Searching for: {search_query} (type: {search_type}) on {site}")
        
        # Perform search based on selected site
        results = []
        
        if site == 'openlibrary':
            results = scraper.search_open_library(search_query, max_results=1)
        elif site == 'barnesandnoble':
            results = barnes_noble_scraper.search_barnes_noble(search_query, max_results=1)
        elif site == 'thriftbooks':
            results = thrift_books_scraper.search_thriftbooks(search_query, max_results=1)
        else:
            # Search both Open Library and Barnes & Noble, combine results
            results = []
        
        # Cache results with timestamp
        cache_key = f"{search_query}_{site}_{datetime.now().strftime('%Y%m%d_%H')}"
        search_cache[cache_key] = {
            'results': results,
            'timestamp': datetime.now().isoformat(),
            'query': search_query,
            'site': site
        }
        
        # Format results for frontend
        formatted_results = format_results_for_frontend(results, site)
        
        # Save results to Excel if requested
        if results and data.get('saveToExcel', False):
            try:
                filename = scraper.save_to_excel(results)
                logger.info(f"Results saved to {filename}")
            except Exception as e:
                logger.error(f"Error saving to Excel: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Found {len(results)} books',
            'results': formatted_results,
            'query': search_query,
            'site': site,
            'total_count': len(results)
        })
        
    except Exception as e:
        logger.error(f"Search error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Search failed: {str(e)}'
        }), 500

def format_results_for_frontend(results, site):
    """Format search results for frontend consumption"""
    formatted = []
    
    for book in results:
        # Handle different result formats from different sources
        formatted_book = {
            'title': book.get('Title', 'Unknown Title'),
            'author': book.get('Author', 'Unknown Author'),
            'isbn': book.get('ISBN', 'N/A'),
            'publisher': book.get('Publisher', 'Unknown Publisher'),
            'publish_date': book.get('Publication_Year', 'Unknown'),
            'price': book.get('Price', ''),  
            'format': book.get('Format', ''),
            'site': site,
            'url': book.get('URL', ''),

        }
        formatted.append(formatted_book)
    
    return formatted


@app.route('/api/history')
def get_search_history():
    """Get recent search history"""
    try:
        # Return recent searches from cache (sorted by timestamp)
        history = []
        for key, data in search_cache.items():
            history.append({
                'query': data['query'],
                'site': data['site'],
                'timestamp': data['timestamp'],
                'result_count': len(data['results'])
            })
        
        # Sort by timestamp (most recent first)
        history.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return jsonify({
            'success': True,
            'history': history[:20]  # Return last 20 searches
        })
        
    except Exception as e:
        logger.error(f"Error fetching search history: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to fetch search history: {str(e)}'
        }), 500

@app.route('/api/export', methods=['POST'])
def export_results():
    """Export search results to Excel and return as downloadable file"""
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({
                'success': False,
                'message': 'No results to export'
            }), 400
        
        # Convert frontend format to DataFrame-ready format
        excel_data = []
        for result in results:
            excel_data.append({
                'Title': result.get('title', 'N/A'),
                'Author': result.get('author', 'N/A'),
                'ISBN': result.get('isbn', 'N/A'),
                'Price': result.get('price', 'N/A'),
                'Format': result.get('format', 'N/A'),
                'Publisher': result.get('publisher', 'N/A'),
                'Publish Date': result.get('publish_date', 'N/A'),
                'Platform': result.get('site', 'N/A'),
                'Book URL': result.get('url', 'N/A'),
                'Subjects': ', '.join(result.get('subjects', [])) if isinstance(result.get('subjects'), list) else result.get('subjects', 'N/A')
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'book_search_results_{timestamp}.xlsx'
        
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        temp_file_path = os.path.join(temp_dir, filename)
        
        # Create Excel file with formatting
        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Book Search Results', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Book Search Results']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add header formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        logger.info(f"Excel file created: {temp_file_path}")
        
        # Return the file as download
        return send_file(
            temp_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }), 500

# Alternative version if you want to clean up temp files after sending
@app.route('/api/export-with-cleanup', methods=['POST'])
def export_results_with_cleanup():
    """Export search results to Excel with automatic cleanup"""
    temp_file_path = None
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({
                'success': False,
                'message': 'No results to export'
            }), 400
        
        # Convert frontend format to DataFrame-ready format
        excel_data = []
        for result in results:
            # Handle subjects field (could be list or string)
            subjects = result.get('subjects', [])
            if isinstance(subjects, list):
                subjects_str = ', '.join(subjects[:5])  # Limit to first 5 subjects
            else:
                subjects_str = str(subjects) if subjects else 'N/A'
            
            excel_data.append({
                'Title': result.get('title', 'N/A'),
                'Author': result.get('author', 'N/A'),
                'ISBN': result.get('isbn', 'N/A'),
                'Price': result.get('price', 'N/A'),
                'Format': result.get('format', 'N/A'),
                'Publisher': result.get('publisher', 'N/A'),
                'Publish Date': result.get('publish_date', 'N/A'),
                'Platform': result.get('site', 'N/A'),
                'Book URL': result.get('url', 'N/A'),
                'Subjects': subjects_str
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'book_search_results_{timestamp}.xlsx'
        
        # Create temporary file
        temp_file_path = tempfile.NamedTemporaryFile(
            delete=False, 
            suffix='.xlsx', 
            prefix='book_export_'
        ).name
        
        # Save to Excel with enhanced formatting
        create_formatted_excel(df, temp_file_path)
        
        logger.info(f"Excel file created: {temp_file_path}")
        
        # Custom response function to clean up after sending
        def remove_file(response):
            try:
                os.unlink(temp_file_path)
                logger.info(f"Temporary file removed: {temp_file_path}")
            except Exception as e:
                logger.warning(f"Could not remove temporary file {temp_file_path}: {e}")
            return response
        
        # Send file and register cleanup
        try:
            response = send_file(
                temp_file_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.call_on_close(lambda: remove_file(None))
            return response
        except Exception as e:
            # If send_file fails, clean up immediately
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            raise e
        
    except Exception as e:
        # Clean up on error
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except:
                pass
        
        logger.error(f"Export error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }), 500

def create_formatted_excel(df, file_path):
    """Create a formatted Excel file from DataFrame"""
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Book Search Results', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Book Search Results']
        
        # Import styling modules
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply cell formatting and auto-adjust column widths
        for col_num, column in enumerate(worksheet.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # Calculate max length for column width
            for cell in column:
                try:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    
                    if cell.row > 1:  # Skip header row for length calculation
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width (with reasonable limits)
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            else:
                worksheet.column_dimensions[column_letter].width = 15
        
        # Set row height for better readability
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 20
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Add filters to header row
        worksheet.auto_filter.ref = worksheet.dimensions

# Clean up old temporary files (call this periodically)
def cleanup_old_temp_files():
    """Clean up old temporary export files"""
    try:
        temp_dir = tempfile.gettempdir()
        current_time = datetime.now()
        
        for filename in os.listdir(temp_dir):
            if filename.startswith('book_export_') and filename.endswith('.xlsx'):
                file_path = os.path.join(temp_dir, filename)
                try:
                    # Remove files older than 1 hour
                    file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                    if (current_time - file_time).seconds > 3600:
                        os.unlink(file_path)
                        logger.info(f"Cleaned up old temp file: {file_path}")
                except Exception as e:
                    logger.warning(f"Could not clean up temp file {file_path}: {e}")
    except Exception as e:
        logger.warning(f"Error during temp file cleanup: {e}")

@app.errorhandler(404)
def not_found(error):
    return jsonify({
        'success': False,
        'message': 'Endpoint not found'
    }), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        'success': False,
        'message': 'Internal server error'
    }), 500

def main():
    """Main function to run the Flask web application"""
    logger.info("Starting Book Search Web Application...")
    
    # Check if the scraper is working
    try:
        logger.info("Book scraper initialized successfully")
    except Exception as e:
        logger.warning(f"Book scraper initialization issue: {str(e)}")
    
    # Run the Flask app
    port = int(os.environ.get('PORT', 5000))
    #debug = os.environ.get('DEBUG', 'False').lower() == 'true'
    
    logger.info(f"Starting server on port {port}")
    logger.info("Access the application at: http://localhost:5000")
    
    app.run(host='0.0.0.0', port=port, debug=True)

if __name__ == "__main__":
    atexit.register(cleanup_old_temp_files)
    main()